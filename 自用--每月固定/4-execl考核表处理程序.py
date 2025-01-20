import os
import tkinter as tk
import zipfile
from datetime import datetime
from tkinter import filedialog, scrolledtext

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side


class ExcelProcessor:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("考核表处理程序")
        self.window.geometry("800x600")

        self.select_button = tk.Button(self.window, text="选择积分文件(.xlsx)", command=self.select_file)
        self.select_button.pack(pady=10)

        self.log_area = scrolledtext.ScrolledText(self.window, width=80, height=30)
        self.log_area.pack(pady=10)

    def log(self, message):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_area.insert(tk.END, f"[{current_time}] {message}\n")
        self.log_area.see(tk.END)

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.process_excel(file_path)

    def get_cell_value(self, cell):
        """获取单元格的值，如果是公式则返回计算结果"""
        if cell.value is None:
            return None
        # 如果单元格包含公式，尝试获取其计算结果
        if isinstance(cell.value, str) and cell.value.startswith('='):
            try:
                return cell.value if cell.value is None else cell.value
            except:
                return None
        return cell.value

    def get_merged_cell_value(self, sheet, row, col):
        for merged_range in sheet.merged_cells.ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row \
                    and col >= merged_range.min_col and col <= merged_range.max_col:
                return self.get_cell_value(sheet.cell(merged_range.min_row, merged_range.min_col))
        return self.get_cell_value(sheet.cell(row, col))

    def find_header_row(self, sheet):
        for row_idx in range(1, 10):
            for cell in sheet[row_idx]:
                if cell.value == "KPI":
                    return row_idx
        return None
    def find_header_summary_row(self, sheet):
        for row_idx in range(1, 10):
            for cell in sheet[row_idx]:
                if cell.value == "合计得分":
                    return row_idx
        return None

    def get_fixed_data(self, fixed_sheet, department, indicator):
        department_row = None
        for row in fixed_sheet.iter_rows(min_row=1):
            if row[0].value == department:
                department_row = row[0].row
                break

        if not department_row:
            return None, None

        weight_col = None
        target_col = None
        for cell in fixed_sheet[1]:
            if cell.value == indicator:
                col_idx = cell.column
                weight_col = col_idx
                target_col = col_idx + 1
                break

        if not weight_col or not target_col:
            return None, None

        weight = self.get_cell_value(fixed_sheet.cell(department_row, weight_col))
        target = self.get_cell_value(fixed_sheet.cell(department_row, target_col))

        return weight, target

    def get_actual_values(self, file_path, department, header_row):
        actual_file_path = file_path.replace('积分', '实际值')

        if not os.path.exists(actual_file_path):
            self.log(f"未找到实际值文件: {actual_file_path}")
            return {}

        try:
            actual_wb = openpyxl.load_workbook(actual_file_path, data_only=True)  # 添加data_only=True参数
            actual_sheet = actual_wb['结果']

            department_row = None
            for row in actual_sheet.iter_rows(min_row=header_row + 1):
                if row[0].value == department:
                    department_row = row
                    break

            if not department_row:
                return {}

            actual_values = {}
            for col in range(1, actual_sheet.max_column + 1):
                indicator = actual_sheet.cell(header_row, col).value
                if indicator:
                    value = self.get_cell_value(department_row[col - 1])
                    if value != "--" and value is not None:
                        actual_values[indicator] = value

            return actual_values

        except Exception as e:
            self.log(f"读取实际值文件时发生错误: {str(e)}")
            return {}

    def process_excel(self, file_path):
        saved_files = []
        try:
            self.log(f"开始处理文件: {file_path}")

            wb = openpyxl.load_workbook(file_path, data_only=True)  # 添加data_only=True参数
            result_sheet = wb['结果']
            fixed_sheet = wb['固定数据']

            header_row = self.find_header_row(result_sheet)
            header_summary_row = self.find_header_summary_row(result_sheet)
            if header_row is None:
                self.log("未找到KPI列")
                return
            if header_summary_row is None:
                self.log("未找到合计得分列")
                return

            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "考核表")
            os.makedirs(desktop_path, exist_ok=True)

            for row in result_sheet.iter_rows(min_row=header_row + 1):
                department = row[0].value
                if not department:
                    continue

                kpi_col = None
                for col in range(1, result_sheet.max_column + 1):
                    if result_sheet.cell(header_row, col).value == "KPI":
                        kpi_col = col
                        break
                summary_col = None
                for col in range(1, result_sheet.max_column + 1):
                    if result_sheet.cell(header_summary_row, col).value == "合计得分":
                        summary_col = col
                        break

                if kpi_col and row[kpi_col - 1].value == "--" and row[summary_col - 1].value != 0:
                    actual_values = self.get_actual_values(file_path, department, header_row)

                    new_wb = openpyxl.Workbook()
                    new_ws = new_wb.active

                    headers = ["指标名称", "权重", "目标值", "实际值", "实际得分"]
                    for col, header in enumerate(headers, 1):
                        new_ws.cell(2, col).value = header
                        new_ws.cell(1, col).value = "总分"

                    current_row = 3
                    for col in range(1, result_sheet.max_column + 1):
                        indicator = result_sheet.cell(header_row, col).value
                        if indicator:
                            value = self.get_cell_value(row[col - 1])
                            if value != "--" and value is not None:
                                weight, target = self.get_fixed_data(fixed_sheet, department, indicator)

                                if weight is not None and target is not None:
                                    new_ws.cell(current_row, 1).value = indicator
                                    new_ws.cell(current_row, 2).value = weight
                                    new_ws.cell(current_row, 3).value = target
                                    actual_value = actual_values.get(indicator)
                                    new_ws.cell(current_row, 4).value = actual_value
                                    new_ws.cell(current_row, 5).value = value
                                    current_row += 1

                    for col in range(2, 6):
                        col_letter = get_column_letter(col)
                        new_ws[f'{col_letter}1'] = f'=SUM({col_letter}3:{col_letter}30)'

                    # 在保存文件之前添加样式设置
                    for row in new_ws.iter_rows(min_row=1, max_row=current_row-1, min_col=1, max_col=5):
                        for cell in row:
                            # 设置边框
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                    # 自动调整列宽
                    for column in new_ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2)
                        new_ws.column_dimensions[column_letter].width = adjusted_width

                    file_name = f"{department}考核表.xlsx"
                    save_path = os.path.join(desktop_path, file_name)
                    new_wb.save(save_path)
                    saved_files.append(save_path)
                    self.log(f"已生成考核表: {file_name}")

            self.log("处理完成!")

        except Exception as e:
            self.log(f"发生错误: {str(e)}")

            # 创建ZIP文件
        zip_path = os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"), f"考核表.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in saved_files:
                arcname = os.path.basename(file_path)
                zipf.write(file_path, arcname)

        print(f"已创建ZIP文件: {zip_path}")

    def run(self):
        self.window.mainloop()


if __name__ == "__main__":
    app = ExcelProcessor()
    app.run()