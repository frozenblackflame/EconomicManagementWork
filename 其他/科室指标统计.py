import json
import pandas as pd
from pathlib import Path
import os
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def get_desktop_path():
    """获取桌面路径"""
    return os.path.join(os.path.expanduser("~"), "Desktop")

def read_json_data():
    """读取JSON文件"""
    desktop_path = get_desktop_path()
    json_path = os.path.join(desktop_path, "指标统计结果.json")
    
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def format_worksheet(worksheet):
    """设置工作表格式"""
    # 设置第一行合并单元格并居中
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    worksheet.column_dimensions['B'].width = 25  # 考核指标列
    worksheet.column_dimensions['P'].width = 11  # 全年均值列
    
    # 设置所有单元格居中
    for row in worksheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整其他列的宽度为合适值
    for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        worksheet.column_dimensions[col].width = 8

def read_target_values():
    """读取目标值"""
    target_file = r"C:\Users\biyun\Desktop\work\2024年绩效\2024年11月绩效.xlsx"
    try:
        # 读取固定数据工作表
        df = pd.read_excel(target_file, sheet_name='固定数据')
        
        # 将科室名称列设为索引
        df.set_index(df.columns[0], inplace=True)
        
        # 创建目标值字典
        target_dict = {}
        for dept in df.index:
            target_dict[dept] = {}
            for col_idx in range(len(df.columns)-1):  # -1 是因为第一列是科室名称
                indicator_name = df.columns[col_idx]
                target_value = df.iloc[:, col_idx+1][dept]  # 错位读取
                target_dict[dept][indicator_name] = target_value
                
        return target_dict
    except Exception as e:
        print(f"读取目标值文件时出错: {e}")
        return {}

def create_excel_data(data):
    """处理数据并创建Excel文件"""
    desktop_path = get_desktop_path()
    excel_path = os.path.join(desktop_path, "指标.xlsx")
    
    # 读取目标值
    target_values = read_target_values()
    
    # 创建ExcelWriter对象
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 遍历每个科室
        for dept_name, dept_data in data.items():
            valid_indicators = []
            
            # 获取当前科室的目标值字典
            dept_targets = target_values.get(dept_name, {})
            
            # 遍历科室下的所有指标
            for indicator_name, indicator_data in dept_data.items():
                if indicator_data.get('monthly_data') and len(indicator_data['monthly_data']) > 0:
                    monthly_data = indicator_data['monthly_data']
                    stats = indicator_data['statistics']
                    
                    # 获取目标值
                    target_value = dept_targets.get(indicator_name, '')
                    
                    row_data = {
                        '序号': len(valid_indicators) + 1,
                        '考核指标': indicator_name,
                        '目标值': target_value,  # 设置目标值
                        '1月': monthly_data.get('01月', ''),
                        '2月': monthly_data.get('02月', ''),
                        '3月': monthly_data.get('03月', ''),
                        '4月': monthly_data.get('04月', ''),
                        '5月': monthly_data.get('05月', ''),
                        '6月': monthly_data.get('06月', ''),
                        '7月': monthly_data.get('07月', ''),
                        '8月': monthly_data.get('08月', ''),
                        '9月': monthly_data.get('09月', ''),
                        '10月': monthly_data.get('10月', ''),
                        '11月': monthly_data.get('11月', ''),
                        '12月': monthly_data.get('12月', ''),
                        '全年均值': stats.get('平均值', '')
                    }
                    valid_indicators.append(row_data)
            
            if valid_indicators:
                # 创建DataFrame
                df = pd.DataFrame(valid_indicators)
                
                # 写入Excel，包括标题行
                header = f"2024年{dept_name}指标统计表"
                df.to_excel(writer, sheet_name=dept_name, index=False, startrow=1)
                
                # 获取worksheet对象并写入标题
                worksheet = writer.sheets[dept_name]
                worksheet.cell(row=1, column=1, value=header)
                
                # 设置格式
                format_worksheet(worksheet)

def main():
    # 读取JSON数据
    data = read_json_data()
    
    # 创建Excel文件
    create_excel_data(data)
    
    print("Excel文件已生成到桌面")

if __name__ == "__main__":
    main()
